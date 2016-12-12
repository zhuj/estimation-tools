#!/usr/bin/env python2.7
# -*- coding: utf-8 -*-

import unittest
import estimate

# openpyxl is required for xlsx reading
# doc: http://openpyxl.readthedocs.io/en/default/
# source: https://bitbucket.org/openpyxl/openpyxl/src/
import openpyxl

def _report(name, root, options):
    fn = ".tmp.%s.xlsx" % name
    estimate.Processor(options).report(root, filename=fn)
    wb = openpyxl.load_workbook(filename=fn, read_only=True)
    names = wb.get_sheet_names()
    return wb[names[0]]

def _check_header(ws):
    assert (ws.min_column, ws.min_row) == (1, 1)
    assert ws['A1'].value == 'Task / Subtask'
    assert ws['B1'].value == 'Filter'
    assert ws['C1'].value == ''
    assert ws['D1'].value == 'Comment'
    assert ws['E1'].value == 'Min'
    assert ws['F1'].value == 'Real'
    assert ws['G1'].value == 'Max'
    assert ws['H1'].value == ''
    assert ws['I1'].value == 'Avg'
    assert ws['J1'].value == 'SD'
    assert ws['K1'].value == 'Sq'
    return 2

def _check_footer(ws):
    row = ws.max_row

    # max
    assert row > 1
    assert str(ws['A%s' % row].value).startswith('Max')
    row -= 1

    # min
    assert row > 1
    assert str(ws['A%s' % row].value).startswith('Min')
    row -= 1

    # kappa
    assert row > 1
    assert str(ws['A%s' % row].value).startswith('K')
    row -= 1

    # standard deviation
    assert row > 1
    assert str(ws['A%s' % row].value).startswith('Standard deviation')
    row -= 1

    # empty
    assert row > 1
    assert str(ws['A%s' % row].value or '') == ''
    row -= 1

    # roles (if exists)
    assert row > 1
    while str(ws['A%s' % row].value).startswith(' - '):
        row -= 1
        assert row > 1

    # total
    assert row > 1
    assert str(ws['A%s' % row].value).startswith('Total')
    row -= 1

    # empty
    assert row > 1
    assert str(ws['A%s' % row].value or '') == ''
    row -= 1

    # return it
    return row

def _check_header_and_footer(ws):
    data_min = _check_header(ws)
    data_max = _check_footer(ws)
    assert data_max > data_min
    return (data_min, data_max)

class ReportTestCase(unittest.TestCase):
    """Test reports itselves"""

    def test_empty_mindmap(self):
        """just check it doesn't throw an exception"""

        root = estimate.Node(parent=None, title=None)
        _report("simple", root, object())

    def test_no_estimates(self):
        """test for simple case without estimates"""

        root = estimate.Node(parent=None, title=None)
        root.append(estimate.Node(parent=root, title="1st node with no estimate"))
        root.append(estimate.Node(parent=root, title="2nd node with no estimate"))
        root.append(estimate.Node(parent=root, title="3rd node with no estimate"))

        ws = _report("no.estimates", root, object())

        data_min, data_max = _check_header_and_footer(ws)
        assert 2 == data_min
        assert 4 == data_max

        assert ws['A2'].value.strip() == "1st node with no estimate"
        assert ws['A3'].value.strip() == "2nd node with no estimate"
        assert ws['A4'].value.strip() == "3rd node with no estimate"

    def test_estimates_sorting(self):
        """estimation roles should be in the end of the child list for the node"""

        root = estimate.Node(parent=None, title=None)

        n1 = estimate.Node(parent=root, title="1st node")
        n1.estimate(role=None, numbers=(0,0,0))
        root.append(n1)

        n2 = estimate.Node(parent=root, title="2nd node")
        root.append(n2)

        n3 = estimate.Node(parent=root, title="3rd node")
        root.append(n3)

        ws = _report("estimate.sort", root, object())

        data_min, data_max = _check_header_and_footer(ws)
        assert 2 == data_min
        assert 4 == data_max

        assert ws['A2'].value.strip() == "2nd node"
        assert ws['A3'].value.strip() == "3rd node"
        assert ws['A4'].value.strip() == "1st node" # estimation rows should be in the end

