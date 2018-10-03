#!/usr/bin/env python
# -*- coding: utf-8 -*-

#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
#
# Copyright (c) 2016 Viktor A. Danilov
# MIT License
# https://opensource.org/licenses/mit-license.html
#

#
# usage: estimate.py [-h] [options] [-o OUTPUT] filename
# see https://github.com/zhuj/estimation-tools/wiki
#
# pre-requirements:
# * pip install openpyxl (http://openpyxl.readthedocs.io/en/default/, https://bitbucket.org/openpyxl/openpyxl/src/)
#

"""
"""

# magic (hack), also it's possible to use export PYTHONIOENCODING=utf8
try:
    import sys
    reload(sys) # reload makes sys.setdefaultencoding method accessible
    sys.setdefaultencoding('utf-8')
except:
    pass



# magic
def magic_split(s):
    part, tp = '', None
    for c in s:
        d = c.isdigit()
        if (tp is None):
            tp = d
        elif (tp != d):
            if (tp is True): part = int(part)
            yield part
            part = ''
            tp = d
        part += c 

    if (tp is True): part = int(part)
    yield part

def magic_compare(a, b):
    return cmp(
        [ x for x in magic_split(a) ],
        [ x for x in magic_split(b) ]
    )

# estimatuion (contains set of numbers)
class Estimation:
    """Estimation is a set of numbers (min <= real <= max) which describes a range for time requirement."""

    @staticmethod
    def _checktype(numbers):
        from types import ListType, TupleType
        if (not isinstance(numbers, (ListType, TupleType))):
            raise TypeError("Invalid operand type: %s", type(numbers))


    def __init__(self, numbers):
        Estimation._checktype(numbers)
        self._numbers = tuple(numbers[0:3])

    def __str__(self):
        return self.__repr__()

    def __repr__(self):
        return str(self._numbers)

    def __getitem__(self, idx):
        return self._numbers[idx]

    def __add__(self, numbers):
        if (isinstance(numbers, Estimation)):
            numbers = numbers._numbers

        Estimation._checktype(numbers)
        return Estimation(tuple(
            self._numbers[i] + numbers[i]
            for i in xrange(3)
        ))


# node (representation of mindmap xml node with additional properties)
class Node:
    """It represents a node in modified mindmap tree (all estimates, roles and comments are collected as attributes)"""

    def __set_parent(self, parent, deep=True):
        if (parent is None):
            self._parent = None
            self._level = -1
        else:
            self._parent = parent
            self._level = 1 + self._parent.level()

        # refresh child nodes, if required
        if (deep and len(self._childs) > 0):
            for c in self._childs:
                c.__set_parent(self, deep=True)

    def __init__(self, parent, title):
        self.__set_parent(parent, deep=False)
        self._title = (title and title.strip() or "")
        self._role = (self._title) and ((self._title[0] == '(') and (self._title[-1] == ')')) or False
        self._annotations = {}
        self._childs = []
        self._estimations = {}
        self._estimates_cache = None

    def __str__(self):
        return self.__repr__()

    def __repr__(self):
        return '<%s level=%s estimates=%s title="%s">' % (
            self.__class__.__name__,
            self.level(),
            self.estimates(),
            self.title()
        )

    def title(self):
        return self._title

    def is_role(self):
        return self._role

    def level(self):
        return self._level

    def title_with_level(self):
        return (('  ' * self.level()) + self.title())

    def parent(self, level=None):
        if (level is None): return self._parent
        if (self._level <= level): return self
        if (self._parent is None): return self
        return self._parent.parent(level)

    def append(self, node, before=None):
        node.__set_parent(self)
        if (before is not None):
            before = self._childs.index(before)
            self._childs.insert(before, node)
        else:
            self._childs.append(node)
        return node

    def detach(self):
        self._parent._childs.remove(self)
        self.__set_parent(None)
        return self

    def childs(self):
        return self._childs

    def annotation(self, name, value=None):
        l = self._annotations.get(name, [])
        if (value is None): return l
        value = value.strip()
        if value:
            l.append(value)
            self._annotations[name] = l
            return l
        return None

    def mvp_minus(self):
        if (len(self.annotation('mvp-')) > 0): return True
        if (self._parent is not None): return self._parent.mvp_minus()
        return False

    def estimate(self, role, numbers):

        # update source
        val = self._estimations.get(role, None)
        if val is None: val = Estimation(numbers)
        else: val = val + numbers
        self._estimations[role] = val

        # and merge it
        self._estimates_cache = Node._estimates(self._estimations)

    def estimates(self):
        return self._estimates_cache

    @staticmethod
    def _estimates(estimations):
        if (len(estimations) <= 0): return None

        estimations = estimations.values()
        if (not estimations): return None

        return reduce(lambda x,y: x+y, estimations)

    def trace(self, stoppers=[]):
        if (self in stoppers): return [ ]
        if (self._parent is None): return [ self ]
        trace = self._parent.trace(stoppers)
        trace.append(self)
        return trace

    def fill_from(self, node):
        import copy
        self._annotations = copy.copy(node._annotations)
        self._estimations = copy.copy(node._estimations)
        self._estimates_cache = Node._estimates(self._estimations)

    def set_custom(self, name, value, check=True, error=lambda prev: None):
        v = getattr(self, '_custom_' + name, None)
        setattr(self, '_custom_' + name, value)
        if (check and (v is not None)):
            err = error(v)
            if (err is not None): raise err
        return v

    def get_custom(self, name):
        return getattr(self, '_custom_' + name, None)

    def acquire_custom(self, name):
        v = self.get_custom(name)
        if (v is not None): return v
        if (self._parent is not None): return self._parent.acquire_custom(name)
        return None


# color management (just to create color series)
import colorsys
def _hls(h, l, s):
    r, g, b = ( int(0xff * i) for i in colorsys.hls_to_rgb(h, l, s) )
    return "#%02X%02X%02X" % (r, g, b)


# default theme (theme wrapper)
class Theme:
    """It wraps given object and looks into it for values: it will use default value it nothing is found.""" 

    _NO_VALUE = object() # placeholder to distinguish true-None and no-value

    # series of lightness
    _SERIES_L = [ 0x84, 0xc2, 0xe0, 0xe8, 0xf0, 0xf8 ]
    _SERIES_L = [ float(l)/0xff for l in _SERIES_L ]

    # saturation (common for all coloring)
    _SERIES_S = 0.318

    # tone for sections (blue sector)
    _SERIES_H = 0.567
    DEFAULT_SECTION_H,\
    DEFAULT_SECTION_1,\
    DEFAULT_SECTION_2,\
    DEFAULT_SECTION_3,\
    DEFAULT_SECTION_4,\
    DEFAULT_SECTION_5 = [
        _hls(_SERIES_H, l, _SERIES_S) for l in _SERIES_L
    ]

    # tone for total values (red sector)
    _SERIES_H = 0.067
    DEFAULT_FINAL,\
    DEFAULT_TOTAL = [
        _hls(_SERIES_H, l, _SERIES_S) for l in (_SERIES_L[2], _SERIES_L[4])
    ]

    # role font color
    DEFAULT_ROLE = '#7f7f7f'

    # default base style
    DEFAULT_F_DEFAULT = {
        'font_name': 'Arial',
        'font_size': 10,
        'align_vertical': 'top'
    }

    # default header format
    DEFAULT_F_HEADER = lambda self: {
        'font_bold': True,
        'border_bottom_color': '#000000',
        'border_bottom_style': 'thin',
        'align_wrap_text': True
    }

    # default cation format
    DEFAULT_F_CAPTION = {
        'font_bold': True
    }

    # default multiplier format
    DEFAULT_F_MULTIPLIER = {
        'num_format': '0'
    }

    # default number format
    DEFAULT_F_NUMBERS = {
        'num_format': '0.00'
    }

    # default boolean format
    DEFAULT_F_BOOLEAN = {
        'num_format': '0'
    }

    # default number format
    DEFAULT_F_ESTIMATES = lambda self: self._merge_format(self.F_NUMBERS, {
        'font_bold': True
    })

    # default number format
    DEFAULT_F_PERCENTAGE = lambda self: self._merge_format(self.F_NUMBERS, {
        'num_format': '0%'
    })

    # comment row
    DEFAULT_F_COMMENT = {
        'align_wrap_text': True,
        'font_italic': False,
        'font_bold': False
    }

    # roles (temporary) row caption
    DEFAULT_F_ROLE_ROW = lambda self: {
        'font_color': self.ROLE,
        'font_italic': False,
        'font_bold': False
    }

    # 1st level (root) sections
    DEFAULT_F_SECTION_1 = lambda self: {
        'font_bold': True,
        'align_wrap_text': True,
        'fill_color': self.SECTION_1
    }

    # 2nd level sections
    DEFAULT_F_SECTION_2 = lambda self: {
        'font_italic': True,
        'align_wrap_text': True,
        'fill_color': self.SECTION_2
    }

    # 3rd level sections
    DEFAULT_F_SECTION_3 = lambda self: {
        'align_wrap_text': True,
        'fill_color': self.SECTION_3
    }

    # 4th level sections
    DEFAULT_F_SECTION_4 = lambda self: {
        'align_wrap_text': True,
        'fill_color': self.SECTION_4
    }

    # 5th level sections
    DEFAULT_F_SECTION_5 = lambda self: {
        'align_wrap_text': True,
        'fill_color': self.SECTION_5
    }

    # total row
    DEFAULT_F_TOTAL = lambda self: {
        'font_bold': True,
        'fill_color': self.TOTAL
    }

    # final values
    DEFAULT_F_FINAL = lambda self: {
        'font_bold': True,
        'fill_color': self.FINAL
    }

    def __init__(self, obj):
        self._obj = obj

    def __getattr__(self, attr):
        """it looks for the attribute in associated object first, then does the same in the theme object itself"""
        v = getattr(self._obj, attr, Theme._NO_VALUE)
        if (v is Theme._NO_VALUE):
            v = getattr(Theme, 'DEFAULT_' + attr)

        if (callable(v)):
            v = v(self)

        if (type(v) is str):
            if (len(v) == 7 and v[0] == '#'): v = 'FF' + v[1:]

        return v

    def _merge_format(self, *formats):
        """it merges (combines) all formats (dicts) together"""
        result = {}
        for f in formats:
            if (callable(f)): f = f(self)
            if (not f): continue
            result.update(f)
        return { k:v for k,v in result.items() if v is not None }

    def format(self, opts = {}):
        """public method for formats: it extends default format with the given one"""
        return self._merge_format(self.F_DEFAULT, opts)


# format cache
class FormatCache:
    """It wraps current theme and call format registration only for completely new format combination"""

    def __init__(self, format=lambda f:f, register=lambda f:f):
        self._cache = {}
        self._format = format
        self._register = register

    @staticmethod
    def _key(format):
        items = format.items()
        items.sort()
        return tuple(items)

    @staticmethod
    def _merge(*formats):
        result = {}
        for f in formats:
            if (not f): continue
            result.update(f)
        return result

    def get(self, *formats):
        format = FormatCache._merge(*formats)
        key = FormatCache._key(format)
        val = self._cache.get(key, None)
        if (val is None):
            format = self._format(format)
            self._cache[key] = val = self._register(format)

        return val


# openpyxl imports
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.cell import column_index_from_string


# column wrapper
class ColumnWrapper:
    """It wraps column"""

    # helper functions

    @staticmethod
    def cell(c, r):
        return "%s%s" % (c, (1+r)) if c else None

    @staticmethod
    def cells(c, r):
        return "%s:%s" % (
            ColumnWrapper.cell(c, r[0]),
            ColumnWrapper.cell(c, r[1])
        ) if c else None

    @staticmethod
    def _name(index):
        return get_column_letter(index+1)

    @staticmethod
    def _index(name):
        return column_index_from_string(name)-1

    def __init__(self, index, apply_format):
        self._index = index
        self._name = ColumnWrapper._name(index) if (type(index) == int) else None
        self._apply_format = apply_format

    def __str__(self):
        return self._name

    def hide(self, ws, hidden=True):
        if (not self._name): return
        ws.column_dimensions[self._name].hidden = hidden

    def setup(self, ws, width, f):
        if (not self._name): return
        self._apply_format(ws.column_dimensions[self._name], f).width = width

    def _apply_format_cell(self, ws, r, *f):
        if (not self._name): return
        return self._apply_format(ws[ColumnWrapper.cell(self._name, r)], *f)

    def string(self, ws, r, string, *f):
        if (not self._name): return
        self._apply_format_cell(ws, r, *f).value = string

    def number(self, ws, r, number, *f):
        if (not self._name): return
        self._apply_format_cell(ws, r, "F_NUMBERS", *f).value = number

    def formula(self, ws, r, formula, *f):
        if (not self._name): return
        self._apply_format_cell(ws, r, "F_NUMBERS", *f).value = formula

    def boolean(self, ws, r, bool, *f):
        if (not self._name): return
        self._apply_format_cell(ws, r, "F_BOOLEAN", *f).value = (bool and 1 or 0)

    def blank(self, ws, r, *f):
        if (not self._name): return
        self.string(ws, r, '', *f)

    def index(self):
        return self._index

    #
    @staticmethod
    def _columns(apply_format, i=0):
        i = 0
        while True:
            yield ColumnWrapper(
                index=i,
                apply_format=apply_format
            )
            i += 1


# processor helper
class Processor:
    """Helper class which does all transformation work"""

    # options
    OPT_MAIN_FACTORS = 'main_factors'
    OPT_MAIN_FACTORS_VARIANT = 'main_factors_variant'
    OPT_ROLE_FACTORS = 'role_factors'
    OPT_THEME = 'theme'
    OPT_SORTING = 'sorting'
    OPT_P_99 = 'p99'
    OPT_MVP = 'mvp'
    OPT_IDS = 'ids'
    OPT_ROLES = 'roles'
    OPT_VALIDATION = 'validation'
    OPT_FORMULAS = 'formulas'
    OPT_FILTER_VISIBILITY = 'filter_visibility'
    OPT_ARROWS = 'arrows'
    OPT_VOLUMES = 'volumes'
    OPT_VOLUMES_MUL = 'volumes_multiplier'
    OPT_STAGES = 'stages'
    OPT_MODULES = 'modules'
    OPT_HIDE_EMPTY_STRUCTURES = "hide_empty_structures"
    OPT_RISKS = "risks"
    OPT_CORRECTIONS = 'corrections'
    OPT_UNPIVOT_TREE = 'unpivot_tree'
    OPT_SEPARATE_FOOTER = 'separate_footer'

    # regexps patterns
    import re
    RE_ESTIMATE = re.compile("estimate\\s*=(\\s*[0-9.]+\\s*)[/](\\s*[0-9.]+\\s*)[/](\\s*[0-9.]+\\s*)")
    RE_ANNOTATION = re.compile("^[[]([a-z0-9-]+)[]]\\s*(.*)")

    # annotation types
    ANNOTATIONS = {
        'warn': ('comment', '(!) '),
        'idea': ('comment', ''),
        'todo': ('comment', ''),
        'impl': ('comment', ''),

        'guess': ('assumption', ''),       # restriction (inclusive) for the estimation
        'assume': ('assumption', ''),      # restriction (inclusive) for the estimation
        'suppose': ('assumption', ''),     # restriction (inclusive) for the estimation
        'assumption': ('assumption', ''),  # restriction (inclusive) for the estimation

        'oos': ('oos', '*'),               # restriction (exclusive)
        'out of scope': ('oos', '*'),      # restriction (exclusive)
        'out-of-scope': ('oos', '*'),      # restriction (exclusive)

        'mvp-': ('mvp-', '*'), # is not MVP
        'api+': ('api+', '*'), # is a common block (API, Architecture, ...)
        'risk': ('risk', '*'), # is a risk (should not happen, but estimated)

        'stage': ('stage', ''),
        'phase': ('stage', ''),
        'module': ('module', '')
    }

    @staticmethod
    def _loadTheme(theme):
        if (isinstance(theme, str)):
            import importlib
            module = importlib.import_module(theme)
            theme = module.Theme

        if (theme is None): theme = object() # just an empty object
        return Theme(theme)

    #
    @staticmethod
    def _wrap_options(options):

        if (isinstance(options, dict)):
            class objectview:
                def __init__(self, d):
                    self.__dict__ = d
            options = objectview(options)

        return options

    #
    @staticmethod
    def _parse_role_factors(factors):
        if (factors):
            factors = factors.split(',')
            factors = [ x.split(':') for x in factors ]
            factors = { x.strip():float(y.strip()) for x, y in factors }
            return factors
        return None

    #
    @staticmethod
    def _parse_main_factors(factors):
        if (factors):
            factors = factors.split('/')
            factors = [ float(f.strip()) for f in factors ]
            while (len(factors) < 3):
                factors.append(factors[-1])

            if (factors == [1.0, 1.0, 1.0]):
                return None

            return factors
        return None

    #
    def __init__(self, options):
        options = Processor._wrap_options(options)
        self._main_factors = Processor._parse_main_factors(getattr(options, Processor.OPT_MAIN_FACTORS, None))
        self._main_factors_variant = (getattr(options, Processor.OPT_MAIN_FACTORS_VARIANT, "") or "").lower()
        self._role_factors = Processor._parse_role_factors(getattr(options, Processor.OPT_ROLE_FACTORS, None))
        self._theme = Processor._loadTheme(getattr(options, Processor.OPT_THEME, None))
        self._sorting = getattr(options, Processor.OPT_SORTING, False) and True
        self._validation = getattr(options, Processor.OPT_VALIDATION, False) and True
        self._mvp = getattr(options, Processor.OPT_MVP, False) and True
        self._ids = getattr(options, Processor.OPT_IDS, False) and True
        self._p99 = getattr(options, Processor.OPT_P_99, False) and True
        self._roles = getattr(options, Processor.OPT_ROLES, False) and True
        self._formulas = self._roles and (getattr(options, Processor.OPT_FORMULAS, False) and True)
        self._filter_visibility = self._roles and (getattr(options, Processor.OPT_FILTER_VISIBILITY, False) and True)
        self._stages = getattr(options, Processor.OPT_STAGES, False) and True
        self._modules = getattr(options, Processor.OPT_MODULES, False) and True
        self._hide_empty = getattr(options, Processor.OPT_HIDE_EMPTY_STRUCTURES, False) and True
        self._risks = getattr(options, Processor.OPT_RISKS, False) and True
        self._arrows = getattr(options, Processor.OPT_ARROWS, False) and True
        self._volumes = getattr(options, Processor.OPT_VOLUMES, False) and True
        self._volumes_mul = self._volumes and float(getattr(options, Processor.OPT_VOLUMES_MUL, "0.0"))
        self._corrections = Processor._parse_role_factors(getattr(options, Processor.OPT_CORRECTIONS, None))
        self._unpivot_tree = getattr(options, Processor.OPT_UNPIVOT_TREE, 0)
        self._separate_footer = getattr(options, Processor.OPT_SEPARATE_FOOTER, False) and True

        if (self._stages):
            self._mvp = False

        if (self._unpivot_tree > 0):
            self._separate_footer = True
            self._roles = False
            self._formulas = False
            self._filter_visibility = False
            self._validation = False


    @staticmethod
    def _attribute(xmlNode, upperName):
        text = ( v for k, v in xmlNode._get_attributes().items() if k.upper() == upperName )
        return text.next()

    @staticmethod
    def _text(xmlNode):
        return Processor._attribute(xmlNode, 'TEXT')

    #
    @staticmethod
    def _apply_factors(factors, estimates, variant='center'):
        if (factors is None):
            return estimates

        if (variant in ('middle', 'center', '')):
            # option (from-center-to-sides)
            estimates[1] = estimates[1]*factors[1]
            estimates[0] = min(estimates[0]*factors[0], estimates[1])
            estimates[2] = max(estimates[2]*factors[2], estimates[1])
        elif (variant == 'right'):
            # option (from-right-to-left)
            estimates[2] = max(estimates[2]*factors[2], 0)
            estimates[1] = min(estimates[1]*factors[1], estimates[2])
            estimates[0] = min(estimates[0]*factors[0], estimates[1])
        elif (variant == 'left'):
            # option (from-left-to-right)
            estimates[0] = max(estimates[0]*factors[0], 0)
            estimates[1] = max(estimates[1]*factors[1], estimates[0])
            estimates[2] = max(estimates[2]*factors[2], estimates[1])
        else:
            raise "Undefined apply factor variant: %s" % variant


        return estimates


    #
    def _process(self, parent, xmlNodes, node_modifier=lambda node: node):
        required = 0

        xmlNodes = [ n for n in xmlNodes if n.nodeType == n.ELEMENT_NODE ]

        # process edges first
        if (self._arrows):
            xmlEdges = [ n for n in xmlNodes if n.tagName == 'arrowlink' ]
            for xmlEdge in xmlEdges:
                target = Processor._attribute(xmlEdge, 'DESTINATION')
                if (target):
                    parent.annotation("dependency_ids", target)
            del xmlEdges

        xmlNodes = [ n for n in xmlNodes if n.tagName == 'node' ]

        # sort by title
        if (self._sorting):
            xmlNodes = [ (Processor._text(x), x) for x in xmlNodes ]
            xmlNodes.sort(lambda x, y: cmp(x[0], y[0]))
            xmlNodes = [ x for title, x in xmlNodes ]

        # start working with nodes
        for xmlNode in xmlNodes:
            title = Processor._text(xmlNode)
            node_id = Processor._attribute(xmlNode, 'ID')

            # first look at the estimation pattern
            match = Processor.RE_ESTIMATE.match(title)
            if (match):
                estimates = [ float(match.group(x).strip()) for x in (1,2,3) ]
                if (self._main_factors is not None):
                    estimates = Processor._apply_factors(self._main_factors, estimates, self._main_factors_variant)

                if (parent.is_role()):
                    role = parent.title()

                    # extra role modification
                    if (self._role_factors is not None):
                        f = role.strip("()").replace(':','_').lower()
                        f = self._role_factors.get(f, 1.0)
                        estimates = [ e*f for e in estimates ]

                    # store the result into the main node (total estimation node)
                    if (parent.parent().is_role()): raise Exception("Check your document structure: there should not be roles-in-roles")
                    parent.parent().estimate(role, estimates)

                parent.estimate(None, estimates) # (always) set the estimation for node itself
                parent.annotation("node_ids", node_id)
                required = 1
                continue

            # then, try to parse the node as an annotation
            for title_line in title.split('\n'):
                match = Processor.RE_ANNOTATION.match(title_line)
                if (match):
                    prefix, text = [ match.group(x).strip() for x in (1, 2) ]
                    prefix = Processor.ANNOTATIONS.get(prefix, None)
                    if (prefix):
                        k, p = prefix
                        parent.annotation(k, p + text)
                        required = 1

            # then try to parse ids (regardless the parameter value)
            if (title and (title.startswith('{') and title.endswith('}'))):
                id_ids, id_sht = title[1:-1].partition(':')[::2]
                def modifier(child, ids=id_ids.strip(), sht=id_sht.strip()):
                    child.set_custom("id_ids", ids, error=lambda prev: Exception(parent, prev))
                    child.set_custom("id_sht", sht, error=lambda prev: Exception(parent, prev))
                    return child

                if (self._process(parent, xmlNode.childNodes, node_modifier=modifier) is not None):
                    parent.annotation("node_ids", node_id)
                    required = 1

                continue

            # then try to parse volumes (regardless the parameter value)
            if (title and (title.startswith('[[') and title.endswith(']]'))):
                title = title[2:-2]
                parent.set_custom("volume", float(title), error=lambda prev: Exception(parent, prev))
                parent._title += ", %s" % title

                if (self._process(parent, xmlNode.childNodes) is not None):
                    parent.annotation("node_ids", node_id)
                    required = 1

                continue

            # else handle it as a regular node
            node = Node(parent, title)
            node = self._process(node, xmlNode.childNodes)
            if (node is not None):
                node = node_modifier(node)
                node.annotation("node_ids", node_id)
                parent.append(node)
                required = 1

        if (not required): return None
        return parent

    #
    def parse(self, path):

        # parse mm document
        from xml.dom import minidom
        xmldoc = minidom.parse(path)
        xmlmap = xmldoc.getElementsByTagName('map')[0]
        xmlmap = xmlmap.childNodes[0]

        # transform document to nodelist
        root = Node(None, "root")
        root = self._process(root, xmlmap.childNodes)
        return root


    #
    @staticmethod
    def _cleanup_tree(node, transformer=lambda node: node):
        """ it traverse the tree removing unnecessary nodes """

        node = transformer(node)
        if (node is None): return None

        if (node.is_role()): return node
        if (node.estimates() is not None): return node
        if ('*' in node.annotation('oos')): return node

        childs = [ Processor._cleanup_tree(c, transformer) for c in node._childs ]
        childs = [ c for c in childs if c is not None ]

        if ((len(childs) > 0) or (node.level() == 0)):
            node._childs = childs
            return node

        return None

    #
    def transform_stages(self, tree):
        """ it extracts stages (from annotations) and rearrange nodes according their stages """

        # setup default "Stage 0", highlight risks


        for l in tree.childs(): # direct childs of the tree root (first level elements)
            if len(l.annotation('stage')) == 0: 
                l.annotation('stage', '0')

        # let's get started
        lines = Processor._collect(tree)
        lines = [ l for l in lines if (not l.is_role()) ]

        # specify reserved stages for annotated elements
        for l in lines:
            if ("*" in l.annotation("oos")): l.annotation('stage', 'out-of-scope')
            elif ("*" in l.annotation("risk")): l.annotation('stage', 'risks')

        # collect data by annotation and by node title together
        lines = (
            [ ( l, s.strip() ) for l in lines for s in l.annotation('stage') ] + 
            [ ( l, l.title().replace('Stage ', '', 1).strip() ) for l in lines if (l.estimates() is None) and (l.title().startswith('Stage ')) ]
        )

        lines = [ ( l, (sl, ) ) for (l, sl) in lines if (sl) ]

        # introduce node cache (to add auxiliary nodes by paths only once)
        _node_cache = {}
        def _add_node(path, caption=lambda x:x[-1]):
            if (not path): return tree

            v = _node_cache.get(path, None)
            if (v is not None): return v

            r = _add_node(path[:-1], caption)
            e = r.append( Node(r, caption(path)) )
            _node_cache[path] = e
            return e

        # TODO: think about: lines.reverse()
        # TODO: make it backward hierarchy sorted
        # append stages first in a right order
        stages = [ sl for (l, sl) in lines ]
        stages = list(set( stages ))
        stages.sort()
        for sl in stages:
            s = _add_node(path=sl, caption=lambda tr:'Stage %s' % '.'.join([ (x if (x and x[0].isdigit()) else ('"%s"' % x.capitalize())) for x in tr ]))
            s.set_custom("stage", sl)
        del stages

        # full list of stoppers (the tree root and all new stages nodes)
        stoppers = [tree] + _node_cache.values()

        # process nodes
        for l, sl in lines:
            s = _node_cache.get(sl)
            trace = l.trace(stoppers)[:-1]

            for t in trace:
                sl = sl + (t.title(), )
                s = _add_node(sl)
                s.fill_from(t)
                s.set_custom("module", t.acquire_custom("module")) # fix bug # TODO: make it more clear

            s.append(l.detach())

            sl = sl + (l.title(), )
            _node_cache[sl] = l

        # clean up the result and return it
        return Processor._cleanup_tree(
            node=tree,
            transformer=lambda n: n
        )

    #
    def transform_modules(self, tree):
        """ it extracts modules (from annotations) and store them to nodes hierarchy  """

        lines = Processor._collect(tree)
        lines = [ ( l, s.strip() ) for l in lines for s in l.annotation('module') if (not l.is_role()) ]
        for l, s in lines:
            if (not s): continue
            l.set_custom("module", s.lower().capitalize())

        return tree

    #
    def transform_arrows(self, tree):
        """ it extracts dependencies from arrows """

        from collections import defaultdict
        id_to_node = defaultdict(list)

        lines = Processor._collect(tree)
        for line in lines:
            for id in line.annotation("node_ids"):
                id_to_node[id].append(line)

        for line in lines:
            for id in line.annotation("dependency_ids"):
                for dep in id_to_node[id]:
                    line.annotation( 'comment', "Requires: " + " / ".join(n.title() for n in dep.trace([tree])) )

        return tree


    #
    def transform_remove_risks(self, tree):
        """ removes all risks from result estimates """

        def is_risk(node):

            if ("*" in node.annotation('risk')): return True

            stages = [ s.lower() for s in node.annotation('stage') ]
            if ("risks" in stages): return True
            if ("risk" in stages): return True

            return False

        return Processor._cleanup_tree(
            node=tree,
            transformer=lambda n: None if is_risk(n) else n
        )

    #
    def transform_remove_oos(self, tree):
        """ removes all out-of-scope from result estimates """

        def is_oos(node):

            if ("*" in node.annotation('oos')): return True

            stages = [ s.lower() for s in node.annotation('stage') ]
            if ("out-of-scope" in stages): return True
            if ("out of scope" in stages): return True

            return False


        return Processor._cleanup_tree(
            node=tree,
            transformer=lambda n: None if is_oos(n) else n
        )

    #
    def transform(self, tree):
        """ it applies transformations to the tree just before its reporting """
        if (self._modules): tree = self.transform_modules(tree)
        if (self._stages): tree = self.transform_stages(tree)
        if (self._arrows): tree = self.transform_arrows(tree)
        if (not self._risks):
            tree = self.transform_remove_oos(tree)
            tree = self.transform_remove_risks(tree)
        return tree

    #
    @staticmethod
    def _collect(root):
        """ it collects the given hierarchy to list of lines """

        lines = []

        def _collectChilds(node):
            childs = node.childs()
            # TODO: XXX: the following code sorts the list in the node object, each time
            childs.sort(lambda x, y: cmp(
                x.estimates() is not None,
                y.estimates() is not None
            ))
            for n in childs:
                lines.append(n)
                _collectChilds(n)

        _collectChilds(root)
        return [ l for l in lines if l ]

    #
    # XXX: remove
    @staticmethod
    def _apply_format(theme):
        """ build cell-format cache apply function """

        _filter_map = lambda map, prefix: {
            k[len(prefix):]: v for (k, v) in map.items() if k.startswith(prefix)
        }

        _cf_font_cache = FormatCache(
            register=lambda f: f and openpyxl.styles.Font(**f) or None
        )

        _cf_fill_cache = FormatCache(
            register=lambda f: f and openpyxl.styles.PatternFill(fill_type='solid', start_color=f.get('color', None)) or None
        )

        _cf_align_cache = FormatCache(
            register=lambda f: f and openpyxl.styles.Alignment(**f) or None
        )

        _cf_border_cache = FormatCache(
            register=lambda f: f and openpyxl.styles.Border(**{
                side: openpyxl.styles.Side(
                    border_style = map.get("style", None),
                    color = map.get("color", None),
                )
                for side, map in (
                    (side, _filter_map(f, side))
                    for side in ('left', 'right', 'top', 'bottom')
                )
            }) or None
        )

        _cf_cache = FormatCache(
            format=lambda f: theme.format(f),
            register=lambda f: {
                'font': _cf_font_cache.get( _filter_map(f, 'font_') ),
                'fill': _cf_fill_cache.get( _filter_map(f, 'fill_') ),
                'alignment': _cf_align_cache.get( _filter_map(f, 'align_') ),
                'border': _cf_border_cache.get( _filter_map(f, 'border_') ),
                'number_format': f.get('num_format', None)
            }
        )

        def _apply_format(cell, *f):
            f = [ ( getattr(theme, x) if isinstance(x, str) else x ) for x in f ] # lookup strings from theme
            f = _cf_cache.get(*f)
            for x in ('font', 'fill', 'alignment', 'number_format'):
                v = f.get(x, None)
                if v is not None: setattr(cell, x, v)
            return cell

        return _apply_format

    #
    def _report_root(self, root, wb):
        # see https://en.wikipedia.org/wiki/Three-point_estimation

        # first transform tree to lines
        lines = Processor._collect(root)
        if (not lines): return # do nothing with empty document

        # ----------------------------------------------------------- #

        # init format (style) table
        f_caption = self._theme.F_CAPTION
        f_comment = self._theme.F_COMMENT
        f_estimates = self._theme.F_ESTIMATES
        f_percentage = self._theme.F_PERCENTAGE
        f_boolean = self._theme.F_BOOLEAN
        f_role = self._theme.F_ROLE_ROW
        f_total = self._theme.F_TOTAL
        f_final = self._theme.F_FINAL
        f_multiplier = self._theme.F_MULTIPLIER

        # cell format cache
        _apply_format = Processor._apply_format(self._theme)
        NOCL = ColumnWrapper(index=None, apply_format=None)

        # helper functions
        cell  = ColumnWrapper.cell
        cells = ColumnWrapper.cells

        def _hide_row(ws, r, hidden=True):
            ws.row_dimensions[1+r].hidden = hidden

        # ----------------------------------------------------------- #

        # --------------
        # esitmates page

        ws_estimates = ws = wb.active
        ws.title = 'Estimates'

        # --------------
        # define columns

        _columns = ColumnWrapper._columns(apply_format = _apply_format)

        LEVELS = []
        if (self._unpivot_tree > 0):
            max_level = 0
            for l in lines:
                if (l.is_role()): continue
                if (l.estimates() is None): continue
                max_level = max(max_level, l.level())

            max_level += 1
            LEVELS = [ (l, 'Level %s' % (1 + l), _columns.next()) for l in range(max_level) ]

            if (self._unpivot_tree >= 2):
                LEVELS.append((-2, "Title (Reduced)", _columns.next()))
                LEVELS.append((-1, "Title (Flat)", _columns.next()))

        BI = _columns.next() if (self._ids) else NOCL                # ?: id: ids
        BS = _columns.next() if (self._ids) else NOCL                # ?: id: short-title
        B0 = _columns.next()                                         # A: caption
        B1 = _columns.next()                                         # B: visibility
        B2 = _columns.next()                                         # C: empty/MVP/Stage

        S0 = _columns.next()                                         # D: structure1 (module, submodule, ...)
        S1 = _columns.next() if (self._unpivot_tree <= 0) else NOCL  # E: structure2 (...)
        S2 = _columns.next() if (self._unpivot_tree <= 0) else NOCL  # F: structure3 (...)
        S3 = _columns.next() if (self._unpivot_tree <= 0) else NOCL  # G: structure4 (...)

        C0 = _columns.next()                                         # H: comment

        E0 = _columns.next()                                         # I: estimate-0
        E1 = _columns.next()                                         # J: estimate-1
        E2 = _columns.next()                                         # K: estimate-2
        E3 = _columns.next() if (self._unpivot_tree <= 0) else NOCL  # L: estimate-separator
        E4 = _columns.next()                                         # M: estimate-wm
        E5 = _columns.next()                                         # N: estimate-st
        E6 = _columns.next()                                         # O: estimate-sq

        R0 = _columns.next()                                         # P: role name

        V0 = _columns.next()                                         # V: volumes (volume)
        V1 = _columns.next()                                         # V: volumes (complexity)

        del _columns

        # -------------
        # setup columns

        for LI, LT, LL in LEVELS:
            LL.setup(ws, width=20, f=self._theme.F_DEFAULT)

        # setup columns: base columns
        BI.setup(ws, width=10, f=self._theme.F_DEFAULT)
        BS.setup(ws, width=20, f=self._theme.F_DEFAULT)
        B0.setup(ws, width=50, f=self._theme.F_DEFAULT)
        B1.setup(ws, width=3,  f=self._theme.F_MULTIPLIER)
        B2.setup(ws, width=8,  f=self._theme.F_COMMENT)
        S0.setup(ws, width=50, f=self._theme.F_COMMENT)
        S1.setup(ws, width=10, f=self._theme.F_COMMENT)
        S2.setup(ws, width=10, f=self._theme.F_COMMENT)
        S3.setup(ws, width=10, f=self._theme.F_COMMENT)
        C0.setup(ws, width=50, f=self._theme.F_COMMENT)
        E0.setup(ws, width=8,  f=self._theme.F_NUMBERS)
        E1.setup(ws, width=8,  f=self._theme.F_NUMBERS)
        E2.setup(ws, width=8,  f=self._theme.F_NUMBERS)
        E3.setup(ws, width=4,  f=self._theme.F_DEFAULT)
        E4.setup(ws, width=8,  f=self._theme.F_NUMBERS)
        E5.setup(ws, width=8,  f=self._theme.F_NUMBERS)
        E6.setup(ws, width=8,  f=self._theme.F_NUMBERS)
        R0.setup(ws, width=3,  f=self._theme.F_DEFAULT)
        V0.setup(ws, width=8,  f=self._theme.F_NUMBERS)
        V1.setup(ws, width=8,  f=self._theme.F_NUMBERS)

        # hide visibility if required
        if (not self._filter_visibility):
            B1.hide(ws, hidden=True)

        # hide module if required
        if (not self._modules):
            S0.hide(ws, hidden=True)

        # hide structure columns
        S1.hide(ws, hidden=True)
        S2.hide(ws, hidden=True)
        S3.hide(ws, hidden=True)

        # hide role column
        R0.hide(ws, hidden=True)

        # volumes
        if (self._volumes):
            V0.number(ws, 0, self._main_factors[1] if self._main_factors is not None else 1.0)
            V1.number(ws, 0, max(0.0, self._volumes_mul))
        else:
            V0.hide(ws, hidden=True)
            V1.hide(ws, hidden=True)

        # hide ids_sht for unpivot >= v2
        if (self._unpivot_tree >= 2):
            BS.hide(ws, hidden=True)

        # start rows
        row = 0

        # ----------------------
        # setup header (row = 1)
        f_header = self._theme.F_HEADER

        # header: levels
        for LI, LT, LL in LEVELS:
            LL.string(ws, row, LT, f_header) # LL: level

        # header: fixed columns
        BI.string(ws, row, 'IDs', f_header)             # BI: ID: IDs
        BS.string(ws, row, 'Title (Short)', f_header)   # BS: ID: Short title
        B0.string(ws, row, 'Task / Subtask', f_header)  # B0: caption
        B1.string(ws, row, 'Filter', f_header)          # B1: visibility
        B2.string(ws, row, '', f_header)                # B2: empty/MVP/Stage
        S0.string(ws, row, 'Module', f_header)          # S0: structure (module, submodule, ....)
        S1.string(ws, row, '', f_header)                # S1: structure
        S2.string(ws, row, '', f_header)                # S2: structure
        S3.string(ws, row, '', f_header)                # S3: structure
        C0.string(ws, row, 'Comment', f_header)         # C0: comment
        E0.string(ws, row, 'Min', f_header)             # E0: estimate
        E1.string(ws, row, 'Real', f_header)            # E1: estimate
        E2.string(ws, row, 'Max', f_header)             # E2: estimate
        E3.string(ws, row, '', f_header)                # E3: empty
        E4.string(ws, row, 'Avg', f_header)             # E4: weighted mean
        E5.string(ws, row, 'SD', f_header)              # E5: standard deviation
        E6.string(ws, row, 'Sq', f_header)              # E6: squared deviation
        R0.string(ws, row, 'Role', f_header)            # R0: role name

        if (self._mvp):
            B2.string(ws, row, 'MVP', f_header)         # B2: MVP
        elif (self._stages):
            B2.string(ws, row, 'Stage', f_header)       # B2: Stage

        # ------------------------
        # prepare data validation

        dv_mul_list = []
        dv_mvp_list = []
        dv_mvp_empty = []
        if (self._validation):

            AZ = ColumnWrapper(index=ColumnWrapper._index("AZ"), apply_format=_apply_format)

            # validation values for B (multiplier) and C (mvp)
            AZ.hide(ws, hidden=True)            # hide AZ
            AZ.blank(ws, 0, f_multiplier)       # AZ1 = empty
            AZ.number(ws, 1, 0, f_multiplier)   # AZ2 = 0
            AZ.number(ws, 2, 1, f_multiplier)   # AZ3 = 1

            # multiplier
            # XXX: note, the validation here is general only (for the whole column)
            # XXX: the idea is that user wants to control which lines are raw data source and which are representation only
            dv_mul_list = openpyxl.worksheet.datavalidation.DataValidation(
                type="list",
                formula1='$AZ$1:$AZ$3', # '', 0, 1
                allow_blank=True
            )
            dv_mul_list.hide_drop_down = True

            # mvp (selecatable)
            dv_mvp_list = openpyxl.worksheet.datavalidation.DataValidation(
                type="list",
                formula1='$AZ$2:$AZ$3', # 0, 1
                allow_blank=False
            )
            dv_mvp_list.hide_drop_down = False

            # mvp (empty only)
            dv_mvp_empty = openpyxl.worksheet.datavalidation.DataValidation(
                type="list",
                formula1="", # none
                allow_blank=True
            )
            dv_mvp_empty.hide_drop_down = True

        # ----------------
        # stages & modules

        def _stage_to_string(node):
            if (self._stages):
                stage = node.acquire_custom('stage')
                if (stage):
                    return '.'.join(str(x) for x in stage)
            return ''

        def _module_to_string(node):
            if (self._modules):
                module = node.acquire_custom('module')
                if (module): return module
            return ''

        # ------------------------
        # start the transformation

        # lines
        stages = set()
        roles = set()
        modules = set()
        row_lines = {}
        for l in lines:

            # then remove role-rows, if it's required
            role = l.is_role()
            if (role and (not self._roles)): continue

            # obtain estimations for the row (total, aggregated from the node and its roles)
            estimates = l.estimates()

            # unpivot mode has to hide all axillary rows
            if (self._unpivot_tree > 0):
                if (estimates is None):
                    continue

            # let's start
            row += 1
            row_lines[id(l)] = row

            # validation over mvp column
            if (self._validation and self._mvp):
                if (estimates is not None):
                    # allow selection for rows with estimations
                    dv_mvp_list.ranges.append(cells(B2, (row, row)))
                else:
                    # no mvp selection is allowed for non-estimate rows
                    dv_mvp_empty.ranges.append(cells(B2, (row, row)))

            # levels, if exists
            for LI, LT, LL in LEVELS:
                if (LI >= 0):
                    if ((self._unpivot_tree < 2) or (LI < l.level())):
                        p = l.parent(level=LI)
                        if (p.level() == LI):
                            LL.string(ws, row, p.title()) # LL: level
                            continue
                LL.blank(ws, row) # LL: level

            # ----------------------
            # special case for roles (they are just info rows)
            if (role):

                # collect roles rows, if roles are required
                roles.add(l.title())

                # set row options (hide it)
                _hide_row(ws, row, hidden=True)

                # fill with values
                B0.string(ws, row, l.title_with_level(), f_role)          # B0 (title)
                B1.number(ws, row, 0, f_role, f_multiplier)               # B1 (visibility/multiplier)
                B2.blank(ws, row, f_role)                                 # B2 (empty/MVP/Stage)
                R0.formula(ws, row, '=TRIM(%s)' % cell(B0, row), f_role)  # R0 (role)
                S0.string(ws, row, '', f_role)                            # S0 (module, submodule, ...)
                S1.string(ws, row, '', f_role)                            # S1 (comment)
                S2.string(ws, row, '', f_role)                            # S2 (comment)
                S3.string(ws, row, '', f_role)                            # S3 (comment)
                C0.string(ws, row, '', f_role)                            # C0 (comment)

                if (estimates is not None):
                    E0.number(ws, row, estimates[0], f_role)              # E0 (estimate)
                    E1.number(ws, row, estimates[1], f_role)              # E1 (estimate)
                    E2.number(ws, row, estimates[2], f_role)              # E2 (estimate)

                # modules
                if (self._modules):
                    module = _module_to_string(l)
                    modules.add(module)
                    f_row_x = None if (estimates is not None) else f_role
                    S0.string(ws, row, module, f_row_x) # S0 (Module, all rows)

                # stage
                # stage
                if (self._stages):
                    stage = _stage_to_string(l)
                    if (not stage): stage = "(?)"
                    B2.string(ws, row, stage, f_role, f_comment) # B2 (Stage, all rows)

                continue

            # ---------------
            # not a role case (visible rows, main data)

            # determine row format (for non-estimation lines only)
            f_row = None
            if (estimates is None):
                try: f_row = getattr(self._theme, 'F_SECTION_%s' % (1 + l.level()))
                except: pass

            # and store them to current line
            l._f_row = f_row

            # multiplier will be used in SUMIFS formulas:
            # '1' means that it's a data for formulas - we have to use it,
            # '' means that it's a role (don't use it in total numbers - use aggregated estimations)
            multiplier = (not role) and (estimates is not None)

            # prepare id-related data
            row_id_ids = l.acquire_custom('id_ids') if (self._unpivot_tree > 0) else l.get_custom('id_ids')
            row_id_sht = l.acquire_custom('id_sht') if (self._unpivot_tree > 0) else l.get_custom('id_sht')

            if (self._unpivot_tree >= 2):
                row_short_title = row_id_sht or l.title()
                if (len(row_short_title) > 40):
                    print "Title is too long:", row_id_ids, row_short_title
                    row_short_title = row_short_title[0:37] + "..."
                LEVELS[-2][-1].string(ws, row, row_short_title, f_row) # short title
                LEVELS[-1][-1].string(ws, row, l.title(), f_row)       # flat title

            elif (self._unpivot_tree > 0):
                row_id_sht = row_id_sht or l.title()

            # let's fill the row with data
            BI.string(ws, row, row_id_ids, f_row, f_comment)                           # BI (IDs: id)
            BS.string(ws, row, row_id_sht, f_row, f_comment)                           # BS (IDs: short title)
            B0.string(ws, row, l.title_with_level(), f_row)                            # B0 (title)
            B1.blank(ws, row, f_row, f_multiplier)                                     # B1 (visibility/multiplier)
            B2.blank(ws, row, f_row)                                                   # B2 (empty/MVP/Stage)
            R0.blank(ws, row, f_row)                                                   # R0 (role)
            S0.string(ws, row, '', f_row)                                              # S0 (module, submodule, ...)
            S1.string(ws, row, '', f_row)                                              # S1 (comment)
            S2.string(ws, row, '', f_row)                                              # S2 (comment)
            S3.string(ws, row, '', f_row)                                              # S3 (comment)
            C0.string(ws, row, ';\n'.join(l.annotation('comment')), f_row, f_comment)  # C0 (comment)
            E0.string(ws, row, '', f_row)                                              # E0 (estimate)
            E1.string(ws, row, '', f_row)                                              # E1 (estimate)
            E2.string(ws, row, '', f_row)                                              # E2 (estimate)
            E3.blank(ws, row, f_row)                                                   # E3 (empty)
            E4.string(ws, row, '', f_row)                                              # E4 (estimate)
            E5.string(ws, row, '', f_row)                                              # E5 (estimate)
            E6.string(ws, row, '', f_row)                                              # E6 (estimate)

            # setup visibility/multiplier
            if (multiplier):
                B1.number(ws, row, 1, f_row, f_multiplier) # B1 (visibility/multiplier)

            # modules
            if (self._modules):
                module = _module_to_string(l)
                modules.add(module)
                f_row_x = f_row if (estimates is not None) else f_role
                S0.string(ws, row, module, f_row_x) # S0 (Module, all rows)

            # setup MVP/Stage for non-role rows
            if (self._mvp):
                if (estimates is not None):
                    mvp = multiplier and self._mvp and True
                    mvp = mvp and (not l.mvp_minus()) and True
                    B2.boolean(ws, row, mvp, f_row, f_comment) # B2 (MVP, estimate rows only)
            elif (self._stages):
                stage = _stage_to_string(l)
                if (not stage): stage = "(?)"
                stages.add(stage)
                f_row_x = f_row if (estimates is not None) else f_role
                B2.string(ws, row, stage, f_row_x, f_comment) # B2 (Stage, all rows)

            # fill estimates
            if (estimates is not None):
                E0.number(ws, row, estimates[0], f_row, f_estimates)               # E0 (estimate)
                E1.number(ws, row, estimates[1], f_row, f_estimates)               # E1 (estimate)
                E2.number(ws, row, estimates[2], f_row, f_estimates)               # E2 (estimate)
                E4.formula(ws, row, '=(%s+4*%s+%s)/6' % (cell(E0, row), cell(E1, row), cell(E2, row)), f_row, f_estimates)  # E1 (weighted mean)
                E5.formula(ws, row, '=(%s-%s)/6' % (cell(E2, row), cell(E0, row)), f_row, f_estimates)                      # E2 (standard deviation)
                E6.formula(ws, row, '=%s*%s' % (cell(E5, row), cell(E5, row)), f_row, f_estimates)                          # E3 (squared deviation)

                if (self._volumes):
                    volume = l.acquire_custom('volume')
                    if (volume):
                        V0.formula(ws, row, '=max(%s, %s*%s*%s)' % (cell(V0, 0), cell(E4, row), float(volume), cell(V1, 0)), f_row)

        # -------------------------------------
        # change estimation numbers to formulas
        if (self._formulas):
            for l in lines:
                if (not l): continue
                if (l.is_role()): continue
                if (not l.childs()): continue

                l_row = row_lines[id(l)]

                # build a formula
                template = "=0"
                for c in l.childs():
                    c_row = row_lines[id(c)]
                    template += "+"+cell("#", c_row)

                # trim the formula
                template = template.replace("=0+", "=")

                # write write to document
                f_row = l._f_row
                if (l.estimates() is not None):
                    E0.formula(ws, l_row, template.replace('#', str(E0)), f_row, f_estimates)  # E0 (estimate)
                    E1.formula(ws, l_row, template.replace('#', str(E1)), f_row, f_estimates)  # E1 (estimate)
                    E2.formula(ws, l_row, template.replace('#', str(E2)), f_row, f_estimates)  # E2 (estimate)
                else:
                    E0.formula(ws, l_row, template.replace('#', str(E0)), f_row, f_role) # E0 (estimate)
                    E1.formula(ws, l_row, template.replace('#', str(E1)), f_row, f_role) # E1 (estimate)
                    E2.formula(ws, l_row, template.replace('#', str(E2)), f_row, f_role) # E2 (estimate)


        # ----------------------------------------------------------- #

        # footer
        row_footer = row+1

        # ----------------------------------
        # calculate ranges & apply filtering
        row_lines = row_lines.values()
        if (len(row_lines) == 0): row_lines = [ row_footer ]
        row_lines = (min(row_lines), max(max(row_lines), row_footer))

        # data validation (multiplier/filter and mvp, if enabled)
        if (self._validation):
            dv_mul_list.ranges.append(cells(B1, row_lines))
            ws.add_data_validation(dv_mul_list)
            if (self._mvp):
                ws.add_data_validation(dv_mvp_empty)
                ws.add_data_validation(dv_mvp_list)

        # set up autofilters (in headers)
        ws.auto_filter.ref = '%s:%s' % (cell('A', 0), cell(E6, row_footer))
        if (self._filter_visibility):
            ws.auto_filter.add_filter_column(B1.index(), ["1"], blank=True) # B1 (visibility/multiplier)

        # ----------------
        # start the footer

        ws_footer = ws_estimates
        if (self._separate_footer):
            ws_footer = ws = wb.create_sheet("Footer")

            # columns
            B0.setup(ws, width=50, f=self._theme.F_DEFAULT)
            C0.setup(ws, width=50, f=self._theme.F_COMMENT)
            E0.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E1.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E2.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E3.setup(ws, width=4,  f=self._theme.F_DEFAULT)
            E4.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E5.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E6.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            R0.setup(ws, width=3,  f=self._theme.F_DEFAULT)

            # hide pivot
            for LI, LT, LL in LEVELS:
                LL.hide(ws, hidden=True)

            # header: fixed columns
            row = 0
            B0.string(ws, row, '', f_header)                # B0: caption
            B1.string(ws, row, '', f_header)                # B1: visibility
            B2.string(ws, row, '', f_header)                # B2: empty/MVP/Stage
            S0.string(ws, row, '', f_header)                # S0: structure (module, submodule, ....)
            S1.string(ws, row, '', f_header)                # S1: structure
            S2.string(ws, row, '', f_header)                # S2: structure
            S3.string(ws, row, '', f_header)                # S3: structure
            C0.string(ws, row, '', f_header)                # C0: comment
            E0.string(ws, row, 'Min', f_header)             # E0: estimate
            E1.string(ws, row, 'Real', f_header)            # E1: estimate
            E2.string(ws, row, 'Max', f_header)             # E2: estimate
            E3.string(ws, row, '', f_header)                # E3: empty
            E4.string(ws, row, 'Avg', f_header)             # E4: weighted mean
            E5.string(ws, row, '', f_header)                # E5: standard deviation
            E6.string(ws, row, '', f_header)                # E6: squared deviation
            R0.string(ws, row, '', f_header)                # R0: role name

            # hide visibility if required
            BI.hide(ws, hidden=True)
            BS.hide(ws, hidden=True)
            B1.hide(ws, hidden=True)

            # hide structure columns
            S1.hide(ws, hidden=True)
            S2.hide(ws, hidden=True)
            S3.hide(ws, hidden=True)

            # hide role column
            R0.hide(ws, hidden=True)

            # footer
            row_footer = row+1

        # ---------
        # total row

        def _total(
                row_total,
                caption='Total',
                row_criteria='%s!%s,"=1"' % (ws_estimates.title, cells(B1, row_lines)),
                prefix="%s!" % ws_estimates.title
        ):
            # total values (it uses row_mul to avoid duuble calculations for roles)
            B0.string(ws, row_total, caption, f_total)                                                                 # B0 (caption)
            B1.string(ws, row_total, '', f_total)                                                                      # B1 (hidden)
            B2.string(ws, row_total, '', f_total)                                                                      # B2
            S0.string(ws, row_total, '', f_total)                                                                      # S0
            S1.string(ws, row_total, '', f_total)                                                                      # S1
            S2.string(ws, row_total, '', f_total)                                                                      # S2
            S3.string(ws, row_total, '', f_total)                                                                      # S3
            C0.string(ws, row_total, '', f_total)                                                                      # C0
            E0.formula(ws, row_total, '=SUMIFS(%s%s,%s)' % (prefix, cells(E0, row_lines), row_criteria), f_total, f_estimates)   # E0 (sum)
            E1.formula(ws, row_total, '=SUMIFS(%s%s,%s)' % (prefix, cells(E1, row_lines), row_criteria), f_total, f_estimates)   # E1 (sum)
            E2.formula(ws, row_total, '=SUMIFS(%s%s,%s)' % (prefix, cells(E2, row_lines), row_criteria), f_total, f_estimates)   # E2 (sum)
            E3.string(ws, row_total, '', f_total)                                                                      # E3
            E4.formula(ws, row_total, '=SUMIFS(%s%s,%s)' % (prefix, cells(E4, row_lines), row_criteria), f_total, f_estimates)   # E4 (sum)
            E5.formula(ws, row_total, '=SUMIFS(%s%s,%s)' % (prefix, cells(E5, row_lines), row_criteria), f_total, f_estimates)   # E5 (sum)
            E6.formula(ws, row_total, '=SUMIFS(%s%s,%s)' % (prefix, cells(E6, row_lines), row_criteria), f_total, f_estimates)   # E6 (sum,sd)
            return row_total

        def _partial(
                row_footer,
                caption,
                total_cell=None,
                row_criteria='%s!%s,"=1"' % (ws_estimates.title, cells(B1, row_lines)),
                prefix="%s!" % ws_estimates.title,
                f_total=f_total,
                sd=False
        ):
            # partial total row
            B0.string(ws, row_footer, caption, f_total)                                                                # B0 (caption)
            B1.string(ws, row_footer, '', f_total)                                                                     # B1 (hidden)
            B2.string(ws, row_footer, '', f_total)                                                                     # B2
            S0.string(ws, row_footer, '', f_total)                                                                     # S0
            S1.string(ws, row_footer, '', f_total)                                                                     # S1
            S2.string(ws, row_footer, '', f_total)                                                                     # S2
            S3.string(ws, row_footer, '', f_total)                                                                     # S3
            C0.string(ws, row_footer, '', f_total)                                                                     # C0
            E0.formula(ws, row_footer, '=SUMIFS(%s%s,%s)' % (prefix, cells(E0, row_lines), row_criteria), f_total, f_estimates)  # E0 (sum)
            E1.formula(ws, row_footer, '=SUMIFS(%s%s,%s)' % (prefix, cells(E1, row_lines), row_criteria), f_total, f_estimates)  # E1 (sum)
            E2.formula(ws, row_footer, '=SUMIFS(%s%s,%s)' % (prefix, cells(E2, row_lines), row_criteria), f_total, f_estimates)  # E2 (sum)
            E4.formula(ws, row_footer, '=(%s+4*%s+%s)/6' % (cell(E0, row_footer), cell(E1, row_footer), cell(E2, row_footer)), f_total, f_estimates) # E4 (local,total)
            if (total_cell):
                E5.formula(ws, row_footer, '=IF(%s>0, %s/%s, 0)' % (total_cell, cell(E4, row_footer), total_cell), f_total, f_percentage)         # E5 (%)
                R0.formula(ws, row_footer, '=IF(%s>0, 1, 0)' % (cell(E5, row_footer)), f_boolean)                          # R0
            if (sd):
                E6.formula(ws, row_footer, '=SUMIFS(%s%s,%s)' % (prefix, cells(E6, row_lines), row_criteria), f_total, f_estimates)           # E6 (sum,sd)
            return row_footer

        # total values (all)
        row_total = row_footer = _total(
            row_total=row_footer + 1,
            caption=(self._mvp and 'Total (All)' or 'Total')
        )

        # total values for each role, if it's required
        if (self._roles):
            # one extra line
            row_footer += 1

            # total (roles) values
            roles = list(roles)
            roles.sort()
            for role in roles:
                row_footer += 1
                row_footer = _partial(
                    row_footer=row_footer,
                    caption='  - %s' % role.strip('()'),
                    row_criteria='''%s!%s,"=0",%s!%s,"%s"''' % (
                        ws_estimates.title, cells(B1, row_lines),
                        ws_estimates.title, cells(R0, row_lines), role
                    ),
                    total_cell='%s!%s' % (
                        ws_footer.title, cell(E4, row_total)
                    )
                )

        if (self._mvp):
            # one extra line
            row_footer += 1

            # total (mvp) values
            # next calculation will use this row as a basis
            row_total = row_footer = _total(
                row_total=row_footer + 1,
                caption='Total (MVP)',
                row_criteria='%s!%s,"=1",%s!%s,"=1"' % (
                    ws_estimates.title, cells(B1, row_lines),
                    ws_estimates.title, cells(B2, row_lines)
                )
            )

        elif (self._stages):
            # one extra line
            row_footer += 1

            # total (stage) values
            stages = list(stages)
            stages.sort(magic_compare)
            for stage in stages:

                row_footer += 1
                row_footer = _partial(
                    row_footer=row_footer,
                    caption=' - Stage %s' % stage,
                    row_criteria='%s!%s,"=1",%s!%s,"%s"' % (
                        ws_estimates.title, cells(B1, row_lines),
                        ws_estimates.title, cells(B2, row_lines), stage
                    ),
                    total_cell='%s!%s' % (ws_footer.title, cell(E4, row_total))
                )

                if (self._hide_empty):
                    s_lines = [ l for l in lines if _stage_to_string(l) == stage if l.estimates() is not None ]
                    if (not s_lines): _hide_row(ws, row_footer, hidden=True)


        if (self._modules):

            # one extra line
            row_footer += 1

            # total (module) values
            modules = list(modules)
            modules.sort(magic_compare)
            for module in modules:
                row_footer += 1
                row_footer = _partial(
                    row_footer=row_footer,
                    caption='- %s' % (('Module «%s»' % module) if module else "Common"),
                    row_criteria='%s!%s,"=1",%s!%s,"%s"' % (
                        ws_estimates.title, cells(B1, row_lines),
                        ws_estimates.title, cells(S0, row_lines), module
                    ),
                    total_cell='%s!%s' % (ws_footer.title, cell(E4, row_total))
                )

                if (self._hide_empty):
                    m_lines = [ l for l in lines if _module_to_string(l) == module if l.estimates() is not None ]
                    if (not m_lines): _hide_row(ws, row_footer, hidden=True)

        # one extra line
        row_footer += 1

        # sigma: standard deviation
        def _sigma(row_sigma, row_total):
            B0.string(ws, row_sigma, 'Standard deviation', f_caption)                     # B0 (caption)
            B2.formula(ws, row_sigma, '=SQRT(%s)' % (cell(E6, row_total)), f_estimates)   # B2 (sigma)
            C0.formula(ws, row_sigma, '=IF(%s>0, %s/%s, 0)' % (cell(E4, row_total), cell(B2, row_sigma), cell(E4, row_total)), f_percentage)     # C0 (sigma,percentage)
            return row_sigma

        row_footer += 1
        row_sigma = _sigma(row_footer, row_total)

        # factors
        if (self._corrections):
            row_footer += 1
            B0.string(ws, row_footer, 'K:', f_caption)                   # B0 (caption)

            kappa_rows = []

            # base factor
            row_footer += 1
            B0.string(ws, row_footer, ' = base', f_caption)              # B0 (caption)
            B2.number(ws, row_footer, 1.0, f_estimates)                  # B2 (kappa)
            kappa_rows.append(row_footer)

            # factors (from options)
            for f, v in self._corrections.items():
                row_footer += 1
                B0.string(ws, row_footer, ' + %s' % f, f_caption)        # B0 (caption)
                B2.number(ws, row_footer, v, f_estimates, f_percentage)  # B2 (kappa)
                kappa_rows.append(row_footer)

            # correction factor (other)
            row_footer += 1
            B0.string(ws, row_footer, ' + other', f_caption)             # B0 (caption)
            B2.number(ws, row_footer, 0.0, f_estimates, f_percentage)    # B2 (kappa)
            kappa_rows.append(row_footer)

            # all together
            kappa_rows = [ min(kappa_rows), max(kappa_rows) ]

            # correction factor (total multiplier)
            row_footer += 1
            B0.string(ws, row_footer, ' * correction', f_caption)        # B0 (caption)
            B2.number(ws, row_footer, 1.0, f_estimates)                  # B2 (kappa)
            kappa_rows.append(row_footer)

            # kappa: correction factor (total, formula)
            row_footer += 1
            row_kappa = row_footer
            B0.string(ws, row_kappa, 'K (total)', f_caption)             # B0 (caption)
            B2.formula(ws, row_kappa, '=SUM(%s)*%s' % (cells(B2, kappa_rows[0:2]), cell(B2, kappa_rows[2])), f_estimates)  # B2 (kappa)
            del kappa_rows
        else:
            # kappa: correction factor
            row_footer += 1
            row_kappa = row_footer
            B0.string(ws, row_kappa, 'K', f_caption)     # B0 (caption)
            B2.number(ws, row_kappa, 1.5, f_estimates)  # B2 (kappa)

        # empty line
        row_footer += 1

        # Min/Max (P=95/99%)
        def _final(
                row_footer,
                sign,
                row_total=row_total,
                row_sigma=row_sigma,
                caption="",
                prefix="%s!" % ws_footer.title
        ):

            p_title = ""
            p_multiplier = 1

            # TODO: make an option
            # a little cheat here:
            # - we should increase estimation (for max case) - it reduces a risk be out of the estimation,
            # - but we should not decrease it (for min case) - what the point to do it so precisely?
            if (sign == '+'):
                if (self._p99):
                    # P=99%, super precision
                    p_title = "P=99%"
                    p_multiplier = 3
                else:
                    # P=95%, regular precision
                    p_title = "P=95%"
                    p_multiplier = 2

            if (self._mvp):
                p_title = "MVP, %s" % p_title if (p_title) else "MVP"

            caption += { '-': 'Min', '+': 'Max' }[sign]
            if p_title: caption += ' (%s)' % p_title

            B0.string(ws, row_footer, caption, f_total)               # B0 (caption)
            B1.string(ws, row_footer, '', f_total)                    # B1
            B2.formula(ws, row_footer, '=%s%s%s*%s' % (cell(E4, row_total), sign, p_multiplier, cell(B2, row_sigma)), f_total, f_estimates)  # B2 (min/max)
            S0.string(ws, row_footer, '', f_total)                                                                                    # S0
            S1.string(ws, row_footer, '', f_total)                                                                                    # S1
            S2.string(ws, row_footer, '', f_total)                                                                                    # S2
            S3.string(ws, row_footer, '', f_total)                                                                                    # S3
            C0.formula(ws, row_footer, '=%s*%s%s' % (cell(B2, row_footer), prefix, cell(B2, row_kappa)),  f_final, f_estimates)       # C0 (modified)
            return row_footer

        # Min (P=95/99%)
        row_footer += 1
        row_footer = _final(row_footer, sign="-")

        # Max (P=95/99%)
        row_footer += 1
        row_footer = _final(row_footer, sign="+")

        # ----------------------------------------------------------- #

        # -----------------------
        # Stages & Modules report
        if (self._stages and self._modules and len(stages) > 1 and len(modules) > 1):

            ws_matrix = ws = wb.create_sheet("StagesAndModules")

            # columns
            B0.setup(ws, width=50, f=self._theme.F_DEFAULT)
            C0.setup(ws, width=50, f=self._theme.F_COMMENT)
            E0.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E1.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E2.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E3.setup(ws, width=4,  f=self._theme.F_DEFAULT)
            E4.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E5.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E6.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            R0.setup(ws, width=3,  f=self._theme.F_DEFAULT)

            # hide pivot
            for LI, LT, LL in LEVELS:
                LL.hide(ws, hidden=True)

            # header: fixed columns
            row = 0
            B0.string(ws, row, '', f_header)                # B0: caption
            B1.string(ws, row, '', f_header)                # B1: visibility
            B2.string(ws, row, '', f_header)                # B2: empty/MVP/Stage
            S0.string(ws, row, '', f_header)                # S0: structure (module, submodule, ....)
            S1.string(ws, row, '', f_header)                # S1: structure
            S2.string(ws, row, '', f_header)                # S2: structure
            S3.string(ws, row, '', f_header)                # S3: structure
            C0.string(ws, row, '', f_header)                # C0: comment
            E0.string(ws, row, 'Min', f_header)             # E0: estimate
            E1.string(ws, row, 'Real', f_header)            # E1: estimate
            E2.string(ws, row, 'Max', f_header)             # E2: estimate
            E3.string(ws, row, '', f_header)                # E3: empty
            E4.string(ws, row, 'Avg', f_header)             # E4: weighted mean
            E5.string(ws, row, '%', f_header)               # E5: standard deviation
            E6.string(ws, row, 'Sq', f_header)              # E6: squared deviation

            # hide visibility if required
            BI.hide(ws, hidden=True)
            BS.hide(ws, hidden=True)
            B1.hide(ws, hidden=True)

            # hide structure columns
            S1.hide(ws, hidden=True)
            S2.hide(ws, hidden=True)
            S3.hide(ws, hidden=True)

            # hide role column
            R0.hide(ws, hidden=True)

            # footer
            row_footer = row+1

            for stage in stages:

                s_lines = [ l for l in lines if _stage_to_string(l) == stage and l.estimates() is not None]
                if (self._hide_empty):
                    # TODO: don't remove but hide the block
                    if (not s_lines): continue

                row_footer += 1
                row_stage = row_footer = _partial(
                    row_footer=row_footer,
                    caption='Stage %s' % stage,
                    row_criteria='%s!%s,"=1",%s!%s,"%s"' % (
                        ws_estimates.title, cells(B1, row_lines),
                        ws_estimates.title, cells(B2, row_lines), stage
                    ),
                    total_cell='%s!%s' % (ws_footer.title, cell(E4, row_total)),
                    f_total=f_final,
                    sd=True
                )

                for module in modules:

                    row_footer += 1
                    row_footer = _partial(
                        row_footer=row_footer,
                        caption=' - %s' % (('Module «%s»' % module) if module else "Common"),
                        row_criteria='%s!%s,"=1",%s!%s,"%s",%s!%s,"%s"' % (
                            ws_estimates.title, cells(B1, row_lines),
                            ws_estimates.title, cells(B2, row_lines), stage,
                            ws_estimates.title, cells(S0, row_lines), module
                        ),
                        total_cell='%s!%s' % (ws_matrix.title, cell(E4, row_stage))
                    )

                    if (self._hide_empty):
                        if (not s_lines): _hide_row(ws, row_footer, hidden=True)
                        else:
                            m_lines = [ l for l in s_lines if _module_to_string(l) == module and l.estimates() is not None ]
                            if (not m_lines): _hide_row(ws, row_footer, hidden=True)


                # sigma: standard deviation
                row_footer += 1
                row_sigma = _sigma(row_footer, row_total=row_stage)

                # Min (P=95/99%)
                row_footer += 1
                row_footer = _final(row_footer, row_total=row_stage, row_sigma=row_sigma, sign="-")

                # Max (P=95/99%)
                row_footer += 1
                row_footer = _final(row_footer, row_total=row_stage, row_sigma=row_sigma, sign="+")

                # TODO: filter (make it work in Excel Online)
                ws.auto_filter.ref = '%s:%s' % (cell('A', 0), cell(R0, row_footer))
                ws.auto_filter.add_filter_column(R0.index(), ['1'], blank=True)

                # yet another empty
                row_footer += 1


        # -----------------------
        # Stages & Roles report
        if (self._stages and self._roles and len(stages) > 1 and len(roles) > 1):

            ws_matrix = ws = wb.create_sheet("StagesAndRoles")

            # columns
            B0.setup(ws, width=50, f=self._theme.F_DEFAULT)
            C0.setup(ws, width=50, f=self._theme.F_COMMENT)
            E0.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E1.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E2.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E3.setup(ws, width=4,  f=self._theme.F_DEFAULT)
            E4.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E5.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E6.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            R0.setup(ws, width=3,  f=self._theme.F_DEFAULT)

            # hide pivot
            for LI, LT, LL in LEVELS:
                LL.hide(ws, hidden=True)

            # header: fixed columns
            row = 0
            B0.string(ws, row, '', f_header)                # B0: caption
            B1.string(ws, row, '', f_header)                # B1: visibility
            B2.string(ws, row, '', f_header)                # B2: empty/MVP/Stage
            S0.string(ws, row, '', f_header)                # S0: structure (module, submodule, ....)
            S1.string(ws, row, '', f_header)                # S1: structure
            S2.string(ws, row, '', f_header)                # S2: structure
            S3.string(ws, row, '', f_header)                # S3: structure
            C0.string(ws, row, '', f_header)                # C0: comment
            E0.string(ws, row, 'Min', f_header)             # E0: estimate
            E1.string(ws, row, 'Real', f_header)            # E1: estimate
            E2.string(ws, row, 'Max', f_header)             # E2: estimate
            E3.string(ws, row, '', f_header)                # E3: empty
            E4.string(ws, row, 'Avg', f_header)             # E4: weighted mean
            E5.string(ws, row, '%', f_header)               # E5: standard deviation
            E6.string(ws, row, 'Sq', f_header)              # E6: squared deviation

            # hide visibility if required
            BI.hide(ws, hidden=True)
            BS.hide(ws, hidden=True)
            B1.hide(ws, hidden=True)

            # hide structure columns
            S1.hide(ws, hidden=True)
            S2.hide(ws, hidden=True)
            S3.hide(ws, hidden=True)

            # hide role column
            R0.hide(ws, hidden=True)

            # footer
            row_footer = row+1

            for stage in stages:

                s_lines = [ l for l in lines if _stage_to_string(l) == stage and l.estimates() is not None ]
                if (self._hide_empty):
                    # TODO: don't remove but hide the block
                    if (not s_lines): continue

                row_footer += 1
                row_stage = row_footer = _partial(
                    row_footer=row_footer,
                    caption='Stage %s' % stage,
                    row_criteria='%s!%s,"=1",%s!%s,"%s"' % (
                        ws_estimates.title, cells(B1, row_lines),
                        ws_estimates.title, cells(B2, row_lines), stage
                    ),
                    total_cell='%s!%s' % (ws_footer.title, cell(E4, row_total)),
                    f_total=f_final,
                    sd=True
                )

                for role in roles:

                    row_footer += 1
                    row_footer = _partial(
                        row_footer=row_footer,
                        caption='  - %s' % role.strip('()'),
                        row_criteria='%s!%s,"=0",%s!%s,"%s",%s!%s,"%s"' % (
                            ws_estimates.title, cells(B1, row_lines),
                            ws_estimates.title, cells(B2, row_lines), stage,
                            ws_estimates.title, cells(R0, row_lines), role
                        ),
                        total_cell='%s!%s' % (ws_matrix.title, cell(E4, row_stage))
                    )

                    if (self._hide_empty):
                        if (not s_lines): _hide_row(ws, row_footer, hidden=True)
                        else:
                            r_lines = [ l for l in s_lines if l.is_role() and l.title() == role and l.estimates() is not None ]
                            if (not r_lines): _hide_row(ws, row_footer, hidden=True)

                # sigma: standard deviation
                row_footer += 1
                row_sigma = _sigma(row_footer, row_total=row_stage)

                # Min (P=95/99%)
                row_footer += 1
                row_footer = _final(row_footer, row_total=row_stage, row_sigma=row_sigma, sign="-")

                # Max (P=95/99%)
                row_footer += 1
                row_footer = _final(row_footer, row_total=row_stage, row_sigma=row_sigma, sign="+")

                # TODO: filter (make it work with Excel Online)
                ws.auto_filter.ref = '%s:%s' % (cell('A', 0), cell(R0, row_footer))
                ws.auto_filter.add_filter_column(R0.index(), ['1'], blank=True)

                # yet another empty
                row_footer += 1


        # -----------------------
        # Modules & Roles report
        if (self._modules and self._roles and len(modules) > 1 and len(roles) > 1):

            ws_matrix = ws = wb.create_sheet("ModulesAndRoles")

            # columns
            B0.setup(ws, width=50, f=self._theme.F_DEFAULT)
            C0.setup(ws, width=50, f=self._theme.F_COMMENT)
            E0.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E1.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E2.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E3.setup(ws, width=4,  f=self._theme.F_DEFAULT)
            E4.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E5.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            E6.setup(ws, width=8,  f=self._theme.F_NUMBERS)
            R0.setup(ws, width=3,  f=self._theme.F_DEFAULT)

            # hide pivot
            for LI, LT, LL in LEVELS:
                LL.hide(ws, hidden=True)

            # header: fixed columns
            row = 0
            B0.string(ws, row, '', f_header)                # B0: caption
            B1.string(ws, row, '', f_header)                # B1: visibility
            B2.string(ws, row, '', f_header)                # B2: empty/MVP/Stage
            S0.string(ws, row, '', f_header)                # S0: structure (module, submodule, ....)
            S1.string(ws, row, '', f_header)                # S1: structure
            S2.string(ws, row, '', f_header)                # S2: structure
            S3.string(ws, row, '', f_header)                # S3: structure
            C0.string(ws, row, '', f_header)                # C0: comment
            E0.string(ws, row, 'Min', f_header)             # E0: estimate
            E1.string(ws, row, 'Real', f_header)            # E1: estimate
            E2.string(ws, row, 'Max', f_header)             # E2: estimate
            E3.string(ws, row, '', f_header)                # E3: empty
            E4.string(ws, row, 'Avg', f_header)             # E4: weighted mean
            E5.string(ws, row, '%', f_header)               # E5: standard deviation
            E6.string(ws, row, 'Sq', f_header)              # E6: squared deviation

            # hide visibility if required
            BI.hide(ws, hidden=True)
            BS.hide(ws, hidden=True)
            B1.hide(ws, hidden=True)

            # hide structure columns
            S1.hide(ws, hidden=True)
            S2.hide(ws, hidden=True)
            S3.hide(ws, hidden=True)

            # hide role column
            R0.hide(ws, hidden=True)

            # footer
            row_footer = row+1

            for module in modules:

                m_lines = [ l for l in lines if _module_to_string(l) == module and l.estimates() is not None ]
                if (self._hide_empty):
                    # TODO: don't remove but hide the block
                    if (not m_lines): continue

                row_footer += 1
                row_module = row_footer = _partial(
                    row_footer=row_footer,
                    caption=' - %s' % (('Module «%s»' % module) if module else "Common"),
                    row_criteria='%s!%s,"=1",%s!%s,"%s"' % (
                        ws_estimates.title, cells(B1, row_lines),
                        ws_estimates.title, cells(S0, row_lines), module
                    ),
                    total_cell='%s!%s' % (ws_footer.title, cell(E4, row_total)),
                    f_total=f_final,
                    sd=True
                )

                for role in roles:

                    row_footer += 1
                    row_footer = _partial(
                        row_footer=row_footer,
                        caption='  - %s' % role.strip('()'),
                        row_criteria='%s!%s,"=0",%s!%s,"%s",%s!%s,"%s"' % (
                            ws_estimates.title, cells(B1, row_lines),
                            ws_estimates.title, cells(S0, row_lines), module,
                            ws_estimates.title, cells(R0, row_lines), role
                        ),
                        total_cell='%s!%s' % (ws_matrix.title, cell(E4, row_module))
                    )

                    if (self._hide_empty):
                        if (not m_lines): _hide_row(ws, row_footer, hidden=True)
                        else:
                            r_lines = [ l for l in m_lines if l.is_role() and l.title() == role and l.estimates() is not None ]
                            if (not r_lines): _hide_row(ws, row_footer, hidden=True)


                # sigma: standard deviation
                row_footer += 1
                row_sigma = _sigma(row_footer, row_total=row_module)

                # Min (P=95/99%)
                row_footer += 1
                row_footer = _final(row_footer, row_total=row_module, row_sigma=row_sigma, sign="-")

                # Max (P=95/99%)
                row_footer += 1
                row_footer = _final(row_footer, row_total=row_module, row_sigma=row_sigma, sign="+")

                # TODO: filter (make it work with Excel Online)
                ws.auto_filter.ref = '%s:%s' % (cell('A', 0), cell(R0, row_footer))
                ws.auto_filter.add_filter_column(R0.index(), ['1'], blank=True)

                # yet another empty
                row_footer += 1

        # --------------------------------
        # assumptions, risks, out of scope

        # TODO: collect out-of-scope separately
        # ws_matrix = ws = wb.create_sheet("Scope")




        return row_lines

    # create a report
    def report(self, root, filename):
        wb = openpyxl.Workbook()
        self._report_root(root, wb)

        if (filename):
            wb.save(filename=filename)
        return wb


# argument parser
def parse_arguments():
    import argparse

    parser = argparse.ArgumentParser(
        description='''Converts freemind estimation to xlsx report.'''
    )

    parser.add_argument(
        '--factor',
        action='store',
        dest=Processor.OPT_MAIN_FACTORS,
        help='''use a given factor for estimates (default = 1.0, it's possible to use O/R/P format)'''
    )

    parser.add_argument(
        '--factor-variant',
        action='store',
        dest=Processor.OPT_MAIN_FACTORS_VARIANT,
        help='''specify behaviour of O/R/P factor format'''
    )

    parser.add_argument(
        '--role-factors',
        action='store',
        dest=Processor.OPT_ROLE_FACTORS,
        help='''use a given factors for roles (in format role1:f1,role2:f2,role3:f3,...)'''
    )

    parser.add_argument(
        '--theme',
        action='store',
        dest=Processor.OPT_THEME,
        help='''use a given .py file as a theme'''
    )

    parser.add_argument(
        '--sort', '-s',
        action='store_true',
        dest=Processor.OPT_SORTING,
        help='''sort children nodes by title'''
    )

    parser.add_argument(
        '--p99',
        action='store_true',
        dest=Processor.OPT_P_99,
        help='''use P=99%% instead of P=95%%'''
    )

    parser.add_argument(
        '--mvp',
        action='store_true',
        dest=Processor.OPT_MVP,
        help='''add Minimum Viable Product (MVP) features'''
    )

    parser.add_argument(
        '--ids',
        action='store_true',
        dest=Processor.OPT_IDS,
        help='''add IDs column'''
    )

    # this option is required for SharePoint
    parser.add_argument(
        '--no-data-validation',
        action='store_false',
        dest=Processor.OPT_VALIDATION,
        help='''don't apply data validation for filter and MVP column (SharePoint fix)'''
    )

    parser.add_argument(
        '--no-roles',
        action='store_false',
        dest=Processor.OPT_ROLES,
        help='''don't provide estimation details for each role'''
    )

    parser.add_argument(
        '--formulas',
        action='store_true',
        dest=Processor.OPT_FORMULAS,
        help='''use formulas for estimation numbers (roles are required)'''
    )

    # this option doesn't work for LibreOffice (due to the bug it doesn't recognize initial filter state)
    parser.add_argument(
        '--filter-visibility',
        action='store_true',
        dest=Processor.OPT_FILTER_VISIBILITY,
        help='''don't hide multiplier/visibility column, use it as a filter instead (doesn't work for LibreOffice)'''
    )

    parser.add_argument(
        '--arrows',
        action='store_true',
        dest=Processor.OPT_ARROWS,
        help='''transformation: handle arrows as dependency indicators'''
    )

    parser.add_argument(
        '--stages',
        action='store_true',
        dest=Processor.OPT_STAGES,
        help='''transformation: use stages markers as nodes/subtree groups (replaces MVP feature)'''
    )

    parser.add_argument(
        '--modules',
        action='store_true',
        dest=Processor.OPT_MODULES,
        help='''transformation: use modules markers as nodes/subtree groups'''
    )

    parser.add_argument(
        '--hide-empty-structures',
        action='store_true',
        dest=Processor.OPT_HIDE_EMPTY_STRUCTURES,
        help='''hide empty structure elements (stages, modules, ...) from the footer'''
    )

    parser.add_argument(
        '--no-risks',
        action='store_false',
        dest=Processor.OPT_RISKS,
        help='''completely remove risks section from the result document'''
    )

    parser.add_argument(
        '--corrections',
        action='store',
        dest=Processor.OPT_CORRECTIONS,
        help='''use extra corrections with default values (in format f1:v1,f2:v2,f3:v3,...)'''
    )

    parser.add_argument(
        '--separate-footer',
        action='store_true',
        dest=Processor.OPT_SEPARATE_FOOTER,
        help='''move footer to a separate worksheet'''
    )

    # experimental
    parser.add_argument(
        '--unpivot-tree-v1',
        action='store_const', const=1,
        dest=Processor.OPT_UNPIVOT_TREE,
        help='''show hierarchy columns (level by level) before the table items (it turns off the header and features as roles, formulas, ...) - V1'''
    )

    # experimental
    parser.add_argument(
        '--unpivot-tree-v2', '--unpivot-tree',
        action='store_const', const=2,
        dest=Processor.OPT_UNPIVOT_TREE,
        help='''show hierarchy columns (level by level) before the table items (it turns off the header and features as roles, formulas, ...) - V2'''
    )

    # experimental
    parser.add_argument(
        '--volumes',
        action='store_true',
        dest=Processor.OPT_VOLUMES,
        help='''transformation: use volumes markers as node/subtree weights'''
    )

    # experimental
    parser.add_argument(
        '--volumes-multiplier',
        action='store',
        dest=Processor.OPT_VOLUMES_MUL,
        help='''transformation: use volumes markers as node/subtree weights (global multiplier, default=0.0)'''
    )

    parser.add_argument(
        '-o',
        action='store',
        dest='output',
        help='''out file name'''
    )

    parser.add_argument(
        'filename',
        help='''a freemind (mindmap) file to be converted'''
    )

    return parser.parse_args()

# let's dance
if __name__ == "__main__":
  def main():

    options = parse_arguments()
    filename = options.filename

    processor = Processor(options)
    root = processor.parse(filename)
    root = processor.transform(root)
    processor.report(root, options.output or (filename + ".xlsx"))

  main()